from flask import Flask, render_template, request, jsonify, Response, send_file
from checker import run_check
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import io
import threading
import uuid
import time

app = Flask(__name__)

# 存储任务进度
tasks = {}


def build_excel(results):
    """生成美化后的 Excel 报告"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "404链接检查报告"

    headers = [
        "问题链接(原始)",
        "最终跳转URL",
        "所在页面",
        "锚文本",
        "锚文本类型",
        "链接类型",
        "HTML位置",
        "HTTP状态码",
        "状态说明",
        "是否重定向",
        "检测时间",
    ]

    # 样式定义
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    status_colors = {
        "404 Not Found": "FFCCCC",
        "重定向后404": "FFB3B3",
        "超时": "FFE5B4",
        "连接失败": "FFE5B4",
    }

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写表头
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 30

    # 写数据
    for row_idx, r in enumerate(results, 2):
        row_data = [
            r.get("full_url", r.get("url", "")),
            r.get("final_url", ""),
            r.get("source_page", ""),
            r.get("anchor_text", ""),
            r.get("anchor_type", ""),
            r.get("link_category", ""),
            r.get("element_position", ""),
            r.get("status_code", ""),
            r.get("status_label", ""),
            "是" if r.get("is_redirect") else "否",
            r.get("check_time", ""),
        ]

        status_label = r.get("status_label", "")
        row_color = status_colors.get(
            status_label, "F9F9F9" if row_idx % 2 == 0 else "FFFFFF"
        )
        row_fill = PatternFill("solid", fgColor=row_color)

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = left if col_idx > 1 else center
            if col_idx == 8:  # 状态码居中
                cell.alignment = center
                if isinstance(val, int) and val == 404:
                    cell.font = Font(bold=True, color="CC0000")

        ws.row_dimensions[row_idx].height = 20

    # 列宽
    col_widths = [45, 45, 40, 30, 12, 16, 12, 12, 16, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    # 汇总 Sheet
    ws2 = wb.create_sheet("汇总统计")
    summary_data = {}
    for r in results:
        label = r.get("status_label", "未知")
        summary_data[label] = summary_data.get(label, 0) + 1

    ws2.append(["状态类型", "数量"])
    for k, v in sorted(summary_data.items(), key=lambda x: -x[1]):
        ws2.append([k, v])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/start", methods=["POST"])
def start_check():
    data = request.json
    urls = [u.strip() for u in data.get("urls", []) if u.strip()]
    if not urls:
        return jsonify({"error": "URL列表为空"}), 400

    task_id = str(uuid.uuid4())
    tasks[task_id] = {
        "status": "running",
        "progress": 0,
        "total": 0,
        "message": "初始化中...",
        "phase": "init",
        "results": [],
        "all_results": [],
    }

    def run():
        def cb(current, total, message, phase):
            tasks[task_id]["progress"] = current
            tasks[task_id]["total"] = total
            tasks[task_id]["message"] = message
            tasks[task_id]["phase"] = phase

        problem, all_r = run_check(urls, progress_callback=cb)
        tasks[task_id]["results"] = problem
        tasks[task_id]["all_results"] = all_r
        tasks[task_id]["status"] = "done"

    t = threading.Thread(target=run, daemon=True)
    t.start()

    return jsonify({"task_id": task_id})


@app.route("/progress/<task_id>")
def progress(task_id):
    def generate():
        while True:
            task = tasks.get(task_id)
            if not task:
                yield f"data: {json.dumps({'error': 'task not found'})}\n\n"
                break
            payload = {
                "status": task["status"],
                "progress": task["progress"],
                "total": task["total"],
                "message": task["message"],
                "phase": task["phase"],
                "problem_count": len(task["results"]),
            }
            yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
            if task["status"] == "done":
                break
            time.sleep(0.8)

    return Response(generate(), mimetype="text/event-stream")


@app.route("/results/<task_id>")
def get_results(task_id):
    task = tasks.get(task_id)
    if not task:
        return jsonify({"error": "not found"}), 404
    return jsonify(
        {
            "problems": task["results"],
            "total_checked": len(task["all_results"]),
            "problem_count": len(task["results"]),
        }
    )


@app.route("/export/<task_id>")
def export_excel(task_id):
    task = tasks.get(task_id)
    if not task or not task["results"]:
        return jsonify({"error": "无数据"}), 404

    output = build_excel(task["results"])
    filename = f"404检查报告_{task_id[:8]}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(debug=True, port=5000, threaded=True)
